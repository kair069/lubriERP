from django.http import HttpResponse
from openpyxl import Workbook
from mi_app.models import Producto

from openpyxl.styles import Font, Alignment, Border, Side

def exportar_productos_excel(request):
    # Crear un libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Definir los encabezados de las columnas
    columnas = [
        "ID", "Nombre del Producto", "Marca", "Categoría", 
        "Tipo de Motor", "Descripción", "Imagen URL", 
        "Código QR", "Precio", "Stock", "Fecha de Modificación"
    ]
    
    # Aplicar la fuente "Times New Roman"
    font = Font(name="Times New Roman", size=12)

    # Aplicar negrita a los encabezados y alinear el texto
    bold_font = Font(name="Times New Roman", size=12, bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    
    # Crear un borde para las celdas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Agregar los encabezados con estilo
    for col_num, column_title in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = bold_font
        cell.alignment = alignment
        cell.border = thin_border

    # Obtener los productos desde la base de datos
    productos = Producto.objects.all()

    # Agregar los datos de los productos
    for row_num, producto in enumerate(productos, 2):  # Empezamos en la fila 2
        fecha_modificacion = producto.fecha_modificacion.replace(tzinfo=None) if producto.fecha_modificacion else None
        row = [
            producto.id, producto.nombre_producto, producto.marca.nombre_marca,
            producto.categoria.nombre_categoria, producto.tipo_motor.tipo_motor,
            producto.descripcion, producto.imagen_url, producto.qr_code,
            producto.precio, producto.stock, fecha_modificacion
        ]
        
        # Agregar cada celda con su formato
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border

    # Ajustar el ancho de las columnas
    for col in range(1, len(columnas) + 1):
        max_length = 0
        for row in range(1, len(productos) + 2):  # Desde la fila de encabezado hasta el final
            cell = ws.cell(row=row, column=col)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[chr(64 + col)].width = adjusted_width

    # Configurar la respuesta HTTP para descargar el archivo Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    wb.save(response)

    return response